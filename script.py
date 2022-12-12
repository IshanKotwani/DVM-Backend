import openpyxl
from openpyxl import Workbook,load_workbook


def populate(genre_name):
    wb = load_workbook(r'C:\Users\LENOVO\Desktop\DVM\Populate.xlsx')
    ws = wb.active

    path_1 = r'C:\Users\LENOVO\Desktop\DVM\Classics.xlsx'
    path_2 = r'C:\Users\LENOVO\Desktop\DVM\Adventure.xlsx'
    path_3 = r'C:\Users\LENOVO\Desktop\DVM\Horror.xlsx'
    path_4 = r'C:\Users\LENOVO\Desktop\DVM\Science Fiction.xlsx'

    wb_1 = load_workbook(path_1)
    ws_1 = wb_1.active
    wb_2 = load_workbook(path_2)
    ws_2 = wb_2.active
    wb_3 = load_workbook(path_3)
    ws_3 = wb_3.active
    wb_4 = load_workbook(path_4)
    ws_4 = wb_4.active

    if genre_name == 1:
        for row in range(2,26):
            Book_name = (ws['A'+str(row)].value)
            Book_ISBN = (ws['B'+str(row)].value)
            Book_Author = (ws['C'+str(row)].value) 
            data = ((4+row), (Book_name), (Book_ISBN), (Book_Author))
            ws_1.append(data)
            wb_1.save(path_1)
                
    elif genre_name == 2:
        for row in range(2,26):
            Book_name = (ws['A'+str(row)].value)
            Book_ISBN = (ws['B'+str(row)].value)
            Book_Author = (ws['C'+str(row)].value) 
            data = ((4+row), (Book_name), (Book_ISBN), (Book_Author))
            ws_2.append(data)
            wb_2.save(path_2)
                
    elif genre_name == 3:
        for row in range(2,26):
            Book_name = (ws['A'+str(row)].value)
            Book_ISBN = (ws['B'+str(row)].value)
            Book_Author = (ws['C'+str(row)].value) 
            data = ((4+row), (Book_name), (Book_ISBN), (Book_Author))
            ws_3.append(data)
            wb_3.save(path_3)
            
    elif genre_name == 4:
        for row in range(2,26):
            Book_name = (ws['A'+str(row)].value)
            Book_ISBN = (ws['B'+str(row)].value)
            Book_Author = (ws['C'+str(row)].value) 
            data = ((4+row), (Book_name), (Book_ISBN), (Book_Author))
            ws_4.append(data)
            wb_4.save(path_4)
    
    
        
            
        


        
    
    
    
    
